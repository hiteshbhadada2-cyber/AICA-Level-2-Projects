from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reconai.models.session import ReconciliationSession
from reconai.models.transaction import (
    MatchStatus,
    TransactionType,
    AuditSeverity,
)

# Professional Financial Color Palette
HEADER_BG = "1E3A8A"       # Deep Corporate Navy
HEADER_FG = "FFFFFF"       # White text
ACCENT_BG = "F1F5F9"       # Slate light tint
BORDER_COLOR = "CBD5E1"    # Slate border
HIGH_SEV_BG = "FEE2E2"     # Light red
HIGH_SEV_FG = "991B1B"     # Dark red
MED_SEV_BG = "FEF3C7"      # Light amber
MED_SEV_FG = "92400E"      # Dark amber
MATCH_BG = "DCFCE7"        # Light green
MATCH_FG = "166534"        # Dark green


class ExcelReportExporter:
    """Exports audit-grade, multi-tab Excel workbooks formatted for Chartered Accountants."""

    def export(self, session: ReconciliationSession, output_path: Union[str, Path]) -> str:
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Lookup maps for quick access
        stmt_map = {s.id: s for s in session.statements}
        ledger_map = {l.id: l for l in session.ledger_entries}

        # 1. Executive Summary & BRS Sheet
        self._build_brs_sheet(wb, session, stmt_map, ledger_map)

        # 2. Matched Transactions Sheet
        self._build_matched_sheet(wb, session, stmt_map, ledger_map)

        # 3. Probable Matches Sheet
        self._build_probable_sheet(wb, session, stmt_map, ledger_map)

        # 4. Unmatched Bank Transactions Sheet
        self._build_unmatched_stmt_sheet(wb, session, stmt_map)

        # 5. Unmatched Ledger Entries Sheet
        self._build_unmatched_ledger_sheet(wb, session, ledger_map)

        # 6. Expense Audit Exceptions Sheet
        self._build_audit_sheet(wb, session)

        # 7. Audit Trail & Integrity Log Sheet
        self._build_audit_trail_sheet(wb, session)

        output_file = Path(output_path)
        wb.save(output_file)
        return str(output_file)

    def _style_header_row(self, ws, row_idx: int, num_cols: int):
        font = Font(name="Segoe UI", size=11, bold=True, color=HEADER_FG)
        fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color=BORDER_COLOR),
            right=Side(style="thin", color=BORDER_COLOR),
            top=Side(style="thin", color=BORDER_COLOR),
            bottom=Side(style="medium", color=HEADER_BG),
        )
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 28

    def _autofit_columns(self, ws, max_col: int = 15):
        for col in ws.iter_cols(min_col=1, max_col=min(ws.max_column, max_col)):
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if "\n" in val:
                    val = max(val.split("\n"), key=len)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 12)

    def _build_brs_sheet(self, wb, session, stmt_map, ledger_map):
        ws = wb.create_sheet(title="BRS & Summary")
        ws.views.sheetView[0].showGridLines = True

        # Title Block
        ws.merge_cells("A1:F1")
        ws["A1"] = f"RECONAI — BANK RECONCILIATION & AUDIT REPORT"
        ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color=HEADER_BG)
        
        ws["A2"] = f"Client: {session.client_name} | Period: {session.period_label} | Generated: {session.updated_at.strftime('%d-%b-%Y %H:%M')}"
        ws["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="475569")

        # Summary KPIs
        ws["A4"] = "Reconciliation Metric"
        ws["B4"] = "Count"
        ws["C4"] = "Amount (INR)"
        self._style_header_row(ws, 4, 3)

        matched = [m for m in session.matches if m.status == MatchStatus.MATCHED]
        probable = [m for m in session.matches if m.status == MatchStatus.PROBABLE]
        unmatched_stmt = [m for m in session.matches if m.status == MatchStatus.UNMATCHED and m.statement_tx_id]
        unmatched_ledger = [m for m in session.matches if m.status == MatchStatus.UNMATCHED and m.ledger_entry_id]

        matched_amt = sum([stmt_map[m.statement_tx_id].amount for m in matched if m.statement_tx_id in stmt_map], Decimal("0.00"))
        probable_amt = sum([stmt_map[m.statement_tx_id].amount for m in probable if m.statement_tx_id in stmt_map], Decimal("0.00"))
        unmatched_stmt_amt = sum([stmt_map[m.statement_tx_id].amount for m in unmatched_stmt if m.statement_tx_id in stmt_map], Decimal("0.00"))
        unmatched_ledger_amt = sum([ledger_map[m.ledger_entry_id].amount for m in unmatched_ledger if m.ledger_entry_id in ledger_map], Decimal("0.00"))

        kpis = [
            ("Fully Matched Transactions", len(matched), float(matched_amt)),
            ("Probable Matches (Review Needed)", len(probable), float(probable_amt)),
            ("Unmatched Bank Entries (Timing / Charges)", len(unmatched_stmt), float(unmatched_stmt_amt)),
            ("Unmatched Ledger Entries (Unpresented / Uncredited)", len(unmatched_ledger), float(unmatched_ledger_amt)),
            ("Total Expense Audit Exceptions Flagged", len(session.audit_flags), "-"),
        ]

        curr_row = 5
        for label, cnt, amt in kpis:
            ws[f"A{curr_row}"] = label
            ws[f"B{curr_row}"] = cnt
            ws[f"C{curr_row}"] = amt
            if isinstance(amt, (int, float)):
                ws[f"C{curr_row}"].number_format = "#,##0.00"
            curr_row += 1

        # BRS Standard Statement
        curr_row += 2
        ws[f"A{curr_row}"] = "BANK RECONCILIATION STATEMENT (BRS)"
        ws[f"A{curr_row}"].font = Font(name="Segoe UI", size=12, bold=True, color=HEADER_BG)
        curr_row += 1

        ws[f"A{curr_row}"] = "Particulars"
        ws[f"B{curr_row}"] = "Ref"
        ws[f"C{curr_row}"] = "Amount (INR)"
        self._style_header_row(ws, curr_row, 3)
        curr_row += 1

        # Calculate BRS Components
        stmt_closing = session.statements[-1].balance if session.statements and session.statements[-1].balance else Decimal("0.00")
        
        ws[f"A{curr_row}"] = "Balance as per Bank Statement (Closing)"
        ws[f"C{curr_row}"] = float(stmt_closing)
        ws[f"C{curr_row}"].number_format = "#,##0.00"
        curr_row += 1

        # Add: Cheques deposited in books but not yet credited by bank
        uncredited = sum([l.amount for l in session.ledger_entries if not l.matched and l.type == TransactionType.CREDIT], Decimal("0.00"))
        ws[f"A{curr_row}"] = "Add: Receipts / Cheques entered in books not yet credited by bank"
        ws[f"C{curr_row}"] = float(uncredited)
        ws[f"C{curr_row}"].number_format = "#,##0.00"
        curr_row += 1

        # Less: Cheques issued in books but not yet presented to bank
        unpresented = sum([l.amount for l in session.ledger_entries if not l.matched and l.type == TransactionType.DEBIT], Decimal("0.00"))
        ws[f"A{curr_row}"] = "Less: Payments / Cheques issued in books not yet presented to bank"
        ws[f"C{curr_row}"] = float(-unpresented)
        ws[f"C{curr_row}"].number_format = "#,##0.00"
        curr_row += 1

        # Adjusted Book Balance
        adjusted_balance = stmt_closing + uncredited - unpresented
        ws[f"A{curr_row}"] = "Estimated Balance as per Books"
        ws[f"A{curr_row}"].font = Font(bold=True)
        ws[f"C{curr_row}"] = float(adjusted_balance)
        ws[f"C{curr_row}"].font = Font(bold=True)
        ws[f"C{curr_row}"].number_format = "#,##0.00"
        curr_row += 2

        # Auditor Remarks & CA Opinion Section
        ws[f"A{curr_row}"] = "AUDITOR REMARKS & PROFESSIONAL OPINION"
        ws[f"A{curr_row}"].font = Font(name="Segoe UI", size=11, bold=True, color=HEADER_BG)
        curr_row += 1

        remarks_text = session.auditor_remarks or "All material items reviewed. Direct bank charges and unpresented payment vouchers have been reconciled."
        ws[f"A{curr_row}"] = "Auditor Remarks:"
        ws[f"A{curr_row}"].font = Font(bold=True)
        ws[f"B{curr_row}"] = remarks_text
        ws.merge_cells(f"B{curr_row}:F{curr_row+1}")
        ws[f"B{curr_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        curr_row += 3

        opinion_text = session.partner_opinion or "In our opinion, the Bank Reconciliation Statement correctly reflects the timing and direct charge differences between the client's books and bank records."
        ws[f"A{curr_row}"] = "CA Partner Opinion:"
        ws[f"A{curr_row}"].font = Font(bold=True)
        ws[f"B{curr_row}"] = opinion_text
        ws.merge_cells(f"B{curr_row}:F{curr_row+1}")
        ws[f"B{curr_row}"].alignment = Alignment(wrap_text=True, vertical="top")

        self._autofit_columns(ws)

    def _build_matched_sheet(self, wb, session, stmt_map, ledger_map):
        ws = wb.create_sheet(title="Matched Items")
        headers = [
            "Match ID", "Confidence", "Rule Applied", "Statement Date", "Statement Narration",
            "Statement Amt (Dr/Cr)", "Ledger Date", "Ledger Particulars", "Voucher No",
            "Ledger Amt (Dr/Cr)", "Auditor Reason",
        ]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        self._style_header_row(ws, 1, len(headers))

        matched = [m for m in session.matches if m.status == MatchStatus.MATCHED]
        row_idx = 2
        for m in matched:
            s = stmt_map.get(m.statement_tx_id)
            l = ledger_map.get(m.ledger_entry_id)
            ws.cell(row=row_idx, column=1, value=m.id[:8])
            ws.cell(row=row_idx, column=2, value=f"{int(m.confidence_score * 100)}%")
            ws.cell(row=row_idx, column=3, value=m.rule_applied)
            ws.cell(row=row_idx, column=4, value=s.date.strftime("%d-%b-%Y") if s else "-")
            ws.cell(row=row_idx, column=5, value=s.description if s else "-")
            ws.cell(row=row_idx, column=6, value=float(s.amount) if s else 0.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=7, value=l.date.strftime("%d-%b-%Y") if l else "-")
            ws.cell(row=row_idx, column=8, value=l.description if l else "-")
            ws.cell(row=row_idx, column=9, value=str(l.source_row_ref) if l else "-")
            ws.cell(row=row_idx, column=10, value=float(l.amount) if l else 0.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=11, value=m.plain_english_reason)
            row_idx += 1

        self._autofit_columns(ws)

    def _build_probable_sheet(self, wb, session, stmt_map, ledger_map):
        ws = wb.create_sheet(title="Probable Matches")
        headers = [
            "Match ID", "Confidence", "Statement Date", "Statement Narration", "Statement Amt",
            "Ledger Date", "Ledger Particulars", "Voucher No", "Ledger Amt", "Reason & Attention Required",
        ]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        self._style_header_row(ws, 1, len(headers))

        probable = [m for m in session.matches if m.status == MatchStatus.PROBABLE]
        row_idx = 2
        for m in probable:
            s = stmt_map.get(m.statement_tx_id)
            l = ledger_map.get(m.ledger_entry_id)
            ws.cell(row=row_idx, column=1, value=m.id[:8])
            ws.cell(row=row_idx, column=2, value=f"{int(m.confidence_score * 100)}%")
            ws.cell(row=row_idx, column=3, value=s.date.strftime("%d-%b-%Y") if s else "-")
            ws.cell(row=row_idx, column=4, value=s.description if s else "-")
            ws.cell(row=row_idx, column=5, value=float(s.amount) if s else 0.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=6, value=l.date.strftime("%d-%b-%Y") if l else "-")
            ws.cell(row=row_idx, column=7, value=l.description if l else "-")
            ws.cell(row=row_idx, column=8, value=str(l.source_row_ref) if l else "-")
            ws.cell(row=row_idx, column=9, value=float(l.amount) if l else 0.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=10, value=m.plain_english_reason)
            row_idx += 1

        self._autofit_columns(ws)

    def _build_unmatched_stmt_sheet(self, wb, session, stmt_map):
        ws = wb.create_sheet(title="Unmatched Statement")
        headers = ["Row Ref", "Date", "Narration / Bank Description", "Amount (INR)", "Type", "Balance", "Diagnostic Cause"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        self._style_header_row(ws, 1, len(headers))

        unmatched_stmt_matches = [m for m in session.matches if m.status == MatchStatus.UNMATCHED and m.statement_tx_id]
        row_idx = 2
        for m in unmatched_stmt_matches:
            s = stmt_map.get(m.statement_tx_id)
            if not s:
                continue
            ws.cell(row=row_idx, column=1, value=str(s.source_row_ref))
            ws.cell(row=row_idx, column=2, value=s.date.strftime("%d-%b-%Y"))
            ws.cell(row=row_idx, column=3, value=s.description)
            ws.cell(row=row_idx, column=4, value=float(s.amount)).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=5, value=s.type.value)
            ws.cell(row=row_idx, column=6, value=float(s.balance) if s.balance else "").number_format = "#,##0.00"
            ws.cell(row=row_idx, column=7, value=m.plain_english_reason)
            row_idx += 1

        self._autofit_columns(ws)

    def _build_unmatched_ledger_sheet(self, wb, session, ledger_map):
        ws = wb.create_sheet(title="Unmatched Books")
        headers = ["Voucher No", "Date", "Particulars / Description", "Party / Account", "Amount (INR)", "Type", "Diagnostic Cause"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        self._style_header_row(ws, 1, len(headers))

        unmatched_ledger_matches = [m for m in session.matches if m.status == MatchStatus.UNMATCHED and m.ledger_entry_id]
        row_idx = 2
        for m in unmatched_ledger_matches:
            l = ledger_map.get(m.ledger_entry_id)
            if not l:
                continue
            ws.cell(row=row_idx, column=1, value=str(l.source_row_ref))
            ws.cell(row=row_idx, column=2, value=l.date.strftime("%d-%b-%Y"))
            ws.cell(row=row_idx, column=3, value=l.description)
            ws.cell(row=row_idx, column=4, value=l.account_name or "-")
            ws.cell(row=row_idx, column=5, value=float(l.amount)).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=6, value=l.type.value)
            ws.cell(row=row_idx, column=7, value=m.plain_english_reason)
            row_idx += 1

        self._autofit_columns(ws)

    def _build_audit_sheet(self, wb, session):
        ws = wb.create_sheet(title="Audit Exceptions")
        headers = ["Voucher / Ref", "Severity", "Category", "Rule Fired", "Auditor Finding & Risk", "Recommended Action"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        self._style_header_row(ws, 1, len(headers))

        row_idx = 2
        for f in session.audit_flags:
            ws.cell(row=row_idx, column=1, value=str(f.source_row_ref))
            
            # Severity with color badge
            sev_cell = ws.cell(row=row_idx, column=2, value=f.severity.value)
            if f.severity == AuditSeverity.HIGH:
                sev_cell.fill = PatternFill(start_color=HIGH_SEV_BG, end_color=HIGH_SEV_BG, fill_type="solid")
                sev_cell.font = Font(name="Segoe UI", bold=True, color=HIGH_SEV_FG)
            elif f.severity == AuditSeverity.MEDIUM:
                sev_cell.fill = PatternFill(start_color=MED_SEV_BG, end_color=MED_SEV_BG, fill_type="solid")
                sev_cell.font = Font(name="Segoe UI", bold=True, color=MED_SEV_FG)

            ws.cell(row=row_idx, column=3, value=f.category.value)
            ws.cell(row=row_idx, column=4, value=f.rule_name)
            ws.cell(row=row_idx, column=5, value=f.plain_english_reason)
            ws.cell(row=row_idx, column=6, value=f.suggested_action)
            row_idx += 1

        self._autofit_columns(ws)

    def _build_audit_trail_sheet(self, wb, session):
        ws = wb.create_sheet(title="Audit Trail Log")
        headers = ["Timestamp", "Action", "Auditor / User", "Details & Traceability"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        self._style_header_row(ws, 1, len(headers))

        row_idx = 2
        for e in session.audit_trail:
            ws.cell(row=row_idx, column=1, value=e.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=2, value=e.action)
            ws.cell(row=row_idx, column=3, value=e.user)
            ws.cell(row=row_idx, column=4, value=e.details)
            row_idx += 1

        self._autofit_columns(ws)
